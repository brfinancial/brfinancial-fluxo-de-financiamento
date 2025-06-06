import streamlit as st
import streamlit_authenticator as stauth
from pathlib import Path
from io import BytesIO
import calendar
from datetime import datetime as dt, time
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Gerador de Planilha de Financiamento",
    layout="centered"
)

# Hardcoded credentials — replace or fetch from st.secrets or .env
USERNAME = "brfinancial"
PASSWORD = "1234"

# Set initial session state
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

def login():
    st.title("Login")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        if username == USERNAME and password == PASSWORD:
            st.session_state.authenticated = True
            # --- Auxiliares de taxa externa ---
            def load_taxas(filepath: str) -> dict:
                taxas = {}
                path = Path(filepath)
                if not path.exists():
                    st.error(f"Arquivo de taxas não encontrado: {filepath}")
                    return taxas
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                blocos = [b.strip() for b in content.split("\n\n") if b.strip()]
                for bloco in blocos:
                    linhas = bloco.splitlines()
                    nome = linhas[0].strip()
                    taxas[nome] = {}
                    for linha in linhas[1:]:
                        if '=' in linha:
                            chave, valor = linha.split('=', 1)
                            try:
                                taxas[nome][chave.strip()] = float(valor.strip())
                            except ValueError:
                                taxas[nome][chave.strip()] = valor.strip()
                return taxas

            # --- Funções de cálculo ---
            HEADER_FILL = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
            DATE_FORMAT = 'dd/mm/yyyy'
            CURRENCY_FORMAT = '"R$" #,##0.00'
            PERCENT_FORMAT = '0.00%'

            def adjust_day(date, preferred_day):
                try:
                    return date.replace(day=preferred_day)
                except ValueError:
                    last = calendar.monthrange(date.year, date.month)[1]
                    return date.replace(day=last)

            def days_in_month(date):
                return calendar.monthrange(date.year, date.month)[1]

            class PaymentTracker:
                def __init__(self, dia_pagamento, taxa_juros):
                    self.last_date = None
                    self.dia = dia_pagamento
                    self.taxa = taxa_juros
                def calculate(self, current_date, saldo):
                    if self.last_date is None:
                        self.last_date = current_date
                        return 0.0, 0, 0.0
                    dias_corridos = (current_date - self.last_date).days
                    taxa_efetiva = ((1+self.taxa) ** (dias_corridos / 30)-1)
                    juros = saldo * ((1+self.taxa) ** (dias_corridos / 30)-1)
                    self.last_date = current_date
                    return juros, dias_corridos, taxa_efetiva

            # --- App Streamlit ---
            def main():
                st.title("Bem-vindo ao gerador de financiamento da Br Financial!")

                # Carrega taxas externas
                taxas_path = 'taxas.txt'
                taxas_por_emp = load_taxas(taxas_path)
                
                # Entradas básicas
                cliente = st.text_input("Qual o nome do cliente?")
                valor_imovel = st.number_input("Qual o valor total do imóvel (R$)", min_value=0.0, step=0.01, format="%.2f")
                dia_pagamento = st.number_input("Qual o dia preferencial de pagamento das parcelas mensais? (1-31)", min_value=1, max_value=31, step=1)

                # Selectbox dinâmico
                empreendimento = st.selectbox("Selecione o empreendimento", options=list(taxas_por_emp.keys()))
                taxas_sel = taxas_por_emp.get(empreendimento, {})
                # Extrai taxas específicas
                TAXA_EMISSAO_CCB = taxas_sel.get('TAXA_EMISSAO_CCB', 0.0)
                TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA = taxas_sel.get('TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA', 0.0)
                TAXA_REGISTRO_IMOVEL = taxas_sel.get('TAXA_REGISTRO_IMOVEL', 0.0)
                TAXA_ESCRITURA_IMOVEL = taxas_sel.get('TAXA_ESCRITURA_IMOVEL',0.0)
                TAXA_SEGURO_PRESTAMISTA_PCT = taxas_sel.get('TAXA_SEGURO_PRESTAMISTA_PCT', 0.0)
                TAXA_INCC = taxas_sel.get('TAXA_INCC', 0.0)
                TAXA_IPCA = taxas_sel.get('TAXA_IPCA', 0.0)
                taxa_pre = taxas_sel.get('taxa_pre', 0.0)
                taxa_pos = taxas_sel.get('taxa_pos', 0.0)
                # extras (percentuais)
                taxas_extras = []
                for chave, val in taxas_sel.items():
                    if chave.endswith('_PCT') and chave not in ['TAXA_SEGURO_PRESTAMISTA_PCT']:
                        periodo = 'pré-entrega da chave' if 'INCC' in chave else 'pós-entrega da chave'
                        taxas_extras.append({'pct': val, 'periodo': periodo})

                # Datas e valores adicionais
                data_base_date = st.date_input("Data-base (data de assinatura do contrato)", value=dt.now().date())
                data_base = dt.combine(data_base_date, time())
                capacidade_pre = st.number_input("Qual a capacidade de pagamento do cliente nas parcelas mensais ANTES da entrega das chaves? (R$)", min_value=0.0, step=0.01)
                data_inicio_pre = dt.combine(st.date_input("Data início dos pagamentos mensais durante a construção (pré-entrega)"), time())
                data_entrega = dt.combine(st.date_input("Data de CONCLUSÃO da obra e entrega das chaves"), time())
                fgts = st.number_input("Valor do FGTS para abatimento do saldo devedor (R$)", min_value=0.0, step=0.01)
                fin_banco = st.number_input("Valor financiado pelo banco (abatimento no saldo devedor) (R$)", min_value=0.0, step=0.01)
                capacidade_pos_antes = st.number_input("Qual a capacidade de pagamento do cliente nas parcelas mensais DEPOIS da entrega das conclusão da obra? (R$)", min_value=0.0, step=0.01)
                val_parcela_banco = st.number_input("Qual o valor da parcela mensal para pagamento do financiamento do banco? (R$)", min_value=0.0, step=0.01)
                capacidade_pos = capacidade_pos_antes - val_parcela_banco

                # Pagamentos não recorrentes
                st.subheader("Pagamentos adicionais às parcelas")
                n_non_rec = st.number_input("Quantos pagamentos adicionais terão? (Caso não haja, deixe zerado)", min_value=0, step=1)
                non_rec = []
                for i in range(int(n_non_rec)):
                    d_date = st.date_input(f"Data do pagamento {i+1}", key=f"nr_d_{i}")
                    d = dt.combine(d_date, time())
                    v = st.number_input(f"Valor pagamento {i+1} (R$)", min_value=0.0, step=0.01, key=f"nr_v_{i}")
                    desc = st.text_input(f"Descrição do pagamento {i+1}", key=f"nr_desc_{i}")
                    assoc = st.checkbox(f"Atribuir a parcela normal do mês?", key=f"nr_assoc_{i}")
                    if assoc:
                        d = adjust_day(d, dia_pagamento)
                    non_rec.append({'data': d, 'tipo': desc, 'valor': v, 'assoc': assoc})

                # Séries semestrais e anuais
                st.subheader("Pagamentos Semestrais")
                n_semi = st.number_input("Quantos pagamentos recorrentes semestrais terão? (Caso não haja, deixe zerado)", min_value=0, step=1)
                semi_series = []
                for i in range(int(n_semi)):
                    d0_date = st.date_input(f"Data das parcelas semestrais {i+1}", key=f"s_d0_{i}")
                    d0 = dt.combine(d0_date, time())
                    v = st.number_input(f"Valor da parcela semestral {i+1} (R$)", min_value=0.0, step=0.01, key=f"s_v_{i}")
                    assoc = st.checkbox(f"Atribuir a parcela normal do mês? {i+1}", key=f"s_assoc_{i}")
                    semi_series.append({'d0': d0, 'v': v, 'assoc': assoc, 'tipo': 'Pagamento Semestral'})

                st.subheader("Pagamentos Anuais")
                n_ann = st.number_input("Quantos pagamentos recorrentes anuais terão? (Caso não haja, deixe zerado)", min_value=0, step=1)
                annual_series = []
                for i in range(int(n_ann)):
                    d0_date = st.date_input(f"Data das parcelas anuais {i+1}", key=f"a_d0_{i}")
                    d0 = dt.combine(d0_date, time())
                    v = st.number_input(f"Valor da parcela anual {i+1} (R$)", min_value=0.0, step=0.01, key=f"a_v_{i}")
                    assoc = st.checkbox(f"Atribuir a parcela normal do mês? {i+1}", key=f"a_assoc_{i}")
                    annual_series.append({'d0': d0, 'v': v, 'assoc': assoc, 'tipo': 'Pagamento Anual'})

                # Geração da planilha
                if st.button("Gerar Planilha"):
                    # --- Agrega séries em non_rec ---
                    for series in semi_series:
                        for n in range(100):
                            d = series['d0'] + relativedelta(months=6 * n)
                            if series['assoc']:
                                d = adjust_day(d, dia_pagamento)
                            non_rec.append({
                                'data': d,
                                'tipo': f"{n+1}ª Parcela Semestral",
                                'valor': series['v'],
                                'assoc': series['assoc']
                            })
                    for series in annual_series:
                        for n in range(100):
                            d = series['d0'] + relativedelta(years=n)
                            if series['assoc']:
                                d = adjust_day(d, dia_pagamento) 
                            non_rec.append({
                                'data': d,
                                'tipo': f"{n+1}ª Parcela Anual",
                                'valor': series['v'],
                                'assoc': series['assoc']
                            })

                    # --- Separa pré e pós entre non_rec ---
                    pre_nr = sorted([e for e in non_rec if e['data'] < data_entrega], key=lambda x: x['data'])
                    post_nr = sorted([e for e in non_rec if e['data'] >= data_entrega], key=lambda x: x['data'])

                    eventos = []
                    saldo = valor_imovel

                    # Data base (assinatura do contrato)
                    eventos.append({
                        'data': data_base,
                        'parcela': '',
                        'tipo': 'Data-Base (assinatura do contrato)',
                        'valor': "-",
                        'juros': 0.0,
                        'dias_corridos': 0,
                        'taxa_efetiva': 0.0,
                        'incc': 0.0,
                        'ipca': 0.0,
                        'taxas_extra': [0.0] * len(taxas_extras),
                        'Total de mudança (R$)': 0.0,
                        'saldo': "-"
                    })

                    tracker_pre = PaymentTracker(dia_pagamento, taxa_pre)
                    tracker_pre.last_date = data_base
                
                    # 1) PRÉ-ENTREGA ------------------------------------------------
                    mc_pre = sc_pre = ac_pre = 0
                    pre_count = 1
                    prev_date = cursor = data_inicio_pre

                    while True:
                        d_evt = adjust_day(cursor, dia_pagamento)
                        if d_evt >= data_entrega:
                            break
                        # não-recorrentes pré não associados entre prev_date e d_evt
                        for ev_nr in [e for e in pre_nr if not e['assoc'] and prev_date < e['data'] < d_evt]:
                            juros, dias_corr, taxa_eff = tracker_pre.calculate(ev_nr['data'], saldo)
                            incc_nr = saldo * TAXA_INCC
                            extras_nr = [saldo * t['pct'] if t['periodo'] in ['pré-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                            total_taxas_nr = sum(extras_nr) + incc_nr
                            abat_nr = ev_nr['valor'] - juros - total_taxas_nr
                            saldo -= abat_nr
                            eventos.append({**ev_nr, 'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                                            'incc': incc_nr, 'ipca': 0.0, 'taxas_extra': extras_nr,
                                            'Total de mudança (R$)': abat_nr, 'saldo': saldo})
                    
                        # Parcela mensal pré-entrega (com associações)
                        juros, dias_corr, taxa_eff = tracker_pre.calculate(d_evt, saldo)
                        incc = saldo * TAXA_INCC
                        extras = [saldo * t['pct'] if t['periodo'] in ['pré-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                        total_taxas = sum(extras) + incc
                        # soma de associações pré
                        # 1) parcela mensal pré (sem associados)
                        valor_parcela = capacidade_pre
                        abat_principal = valor_parcela - juros - total_taxas
                        saldo -= abat_principal
                        eventos.append({
                            'data': d_evt,
                            'parcela': pre_count,
                            'tipo': f"{pre_count}ª Parcela Pré-Entrega",
                            'valor': valor_parcela,
                            'juros': juros,
                            'dias_corridos': dias_corr,
                            'taxa_efetiva': taxa_eff,
                            'incc': incc,
                            'ipca': 0.0,
                            'taxas_extra': extras,
                            'Total de mudança (R$)': abat_principal,
                            'saldo': saldo
                        })

                        # 2) cada pagamento adicional associado em linha própria
                        for ev_assoc in [e for e in pre_nr if e['assoc'] and e['data'] == d_evt]:
                            juros_a, dias_a, txef_a = tracker_pre.calculate(ev_assoc['data'], saldo)
                            incc_a = saldo * TAXA_INCC
                            extras_a = [saldo * t['pct'] for t in taxas_extras if t['periodo'] in ['pré-entrega da chave','ambos']]
                            total_taxas_a = incc_a + sum(extras_a)
                            abat_a = ev_assoc['valor'] - juros_a - total_taxas_a
                            saldo -= abat_a
                            eventos.append({
                                'data': d_evt,
                                'parcela': '',
                                'tipo': ev_assoc['tipo'],
                                'valor': ev_assoc['valor'],
                                'juros': juros_a,
                                'dias_corridos': dias_a,
                                'taxa_efetiva': txef_a,
                                'incc': incc_a,
                                'ipca': 0.0,
                                'taxas_extra': extras_a,
                                'Total de mudança (R$)': abat_a,
                                'saldo': saldo
                            })
                        pre_count += 1
                        prev_date = d_evt
                        cursor += relativedelta(months=1)


                    # 2) ENTREGA ------------------------------------------------------
                    ent = data_entrega
                    zero_extras = [0.0] * len(taxas_extras)
                    # abatimentos
                    for desc, v in [('Abatimento FGTS', fgts), ('Abatimento Fin. Banco', fin_banco)]:
                        saldo -= v
                        eventos.append({'data': ent, 'parcela': '', 'tipo': desc, 'valor': v,
                                        'juros': 0.0, 'dias_corridos': '', 'taxa_efetiva': '',
                                        'incc': 0.0, 'ipca': 0.0, 'taxas_extra': zero_extras,
                                        'Total de mudança (R$)': v, 'saldo': saldo})
                    # taxas de emissão e registro
                    for nome, val in [('Emissão CCB', TAXA_EMISSAO_CCB), ('Alienação Fiduciária', TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA),
                                    ('Registro', TAXA_REGISTRO_IMOVEL), ('Escritura Imóvel', TAXA_ESCRITURA_IMOVEL)]:
                        saldo += val
                    # seguro prestamista
                    fee = saldo * TAXA_SEGURO_PRESTAMISTA_PCT
                    saldo += fee
                    
                    #Data da entrega
                    eventos.append({
                        'data': data_entrega,
                        'parcela': '',
                        'tipo': 'Data da entrega das chaves',
                        'valor': "-",
                        'juros': "-",
                        'dias_corridos': "-",
                        'taxa_efetiva': "-",
                        'incc': "-",
                        'ipca': "-",
                        'taxas_extra': "-",
                        'Total de mudança (R$)': "-",
                        'saldo': saldo
                    })

                    # 3) PÓS-ENTREGA --------------------------------------------------
                    tracker_pos = PaymentTracker(dia_pagamento, taxa_pos)
                    tracker_pos.last_date = data_entrega
                    prev_date = data_entrega
                    cursor = data_entrega
                    post_count = 1
                    parcelas = 1
                    while saldo > 0:
                        d_evt = adjust_day(cursor + relativedelta(months=1), dia_pagamento)
                        # não-recorrentes pós não associados entre prev_date e d_evt
                        for ev_nr in [e for e in post_nr if not e['assoc'] and prev_date < e['data'] < d_evt]:
                            juros, dias_corr, taxa_eff = tracker_pos.calculate(ev_nr['data'], saldo)
                            ipca_nr = saldo * TAXA_IPCA
                            extras_nr = [saldo * t['pct'] if t['periodo'] in ['pós-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                            total_taxas_nr = sum(extras_nr) + ipca_nr
                            abat_nr = ev_nr['valor'] - juros - total_taxas_nr
                            saldo -= abat_nr
                            eventos.append({**ev_nr,'parcela': parcelas, 'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                                            'incc': 0.0, 'ipca': ipca_nr, 'taxas_extra': extras_nr,
                                            'Total de mudança (R$)': abat_nr, 'saldo': saldo})

                        # parcela mensal pós-entrega (com associações)
                        # 1) parcela mensal pós (sem associados)
                        juros, dias_corr, txef = tracker_pos.calculate(d_evt, saldo)
                        ipca = saldo * TAXA_IPCA
                        extras = [saldo * t['pct'] for t in taxas_extras if t['periodo'] in ['pós-entrega da chave','ambos']]
                        abat_princ = capacidade_pos - juros - (ipca + sum(extras))
                        saldo -= abat_princ
                        eventos.append({
                            'data': d_evt,
                            'parcela': post_count,
                            'tipo': f"{post_count}ª Parcela Pós-Entrega",
                            'valor': capacidade_pos,
                            'juros': juros,
                            'dias_corridos': dias_corr,
                            'taxa_efetiva': txef,
                            'incc': 0.0,
                            'ipca': ipca,
                            'taxas_extra': extras,
                            'Total de mudança (R$)': abat_princ,
                            'saldo': saldo
                        })

                        # 2) cada pagamento adicional associado em linha própria
                        for ev_assoc in [e for e in post_nr if e['assoc'] and e['data'] == d_evt]:
                            juros_a, dias_a, txef_a = tracker_pos.calculate(ev_assoc['data'], saldo)
                            ipca_a = saldo * TAXA_IPCA
                            extras_a = [saldo * t['pct'] for t in taxas_extras if t['periodo'] in ['pós-entrega da chave','ambos']]
                            total_taxas_a = ipca_a + sum(extras_a)
                            abat_a = ev_assoc['valor'] - juros_a - total_taxas_a
                            saldo -= abat_a
                            eventos.append({
                                'data': d_evt,
                                'parcela': '',
                                'tipo': ev_assoc['tipo'],
                                'valor': ev_assoc['valor'],
                                'juros': juros_a,
                                'dias_corridos': dias_a,
                                'taxa_efetiva': txef_a,
                                'incc': 0.0,
                                'ipca': ipca_a,
                                'taxas_extra': extras_a,
                                'Total de mudança (R$)': abat_a,
                                'saldo': saldo
                            })
                            
                        post_count += 1
                        parcelas += 1
                        prev_date = d_evt
                        cursor = d_evt

                    # --- Montar planilha ---
                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"Financ-{cliente}"[:31]


                    headers = ["Data","Tipo","Valor Pago (R$)"]
                    for i, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=i, value=h)
                        cell.fill = HEADER_FILL
                        cell.font = Font(bold=True)
                    # linha inicial
                    ws.append(["-", "-", valor_imovel])
                    # eventos
                    for ev in sorted(eventos, key=lambda x: x['data']):
                        row = [ev['data'], ev['tipo'], ev.get('valor', 0)]
                        ws.append(row)
                        
                    # 3) Insere linha em branco
                    ws.append([''] * len(headers))
                    
                    # 4) Insere linha de TOTAIS
                    soma_total = sum(ev['valor'] for ev in eventos if isinstance(ev['valor'], (int, float)))
                    ws.append(['TOTAIS', '', soma_total])
                    totals_row = ws.max_row
                    ws.cell(row=totals_row, column=1).fill = HEADER_FILL
                    ws.cell(row=totals_row, column=1).font = Font(bold=True)

                    # Formatação de colunas
                    for col_idx, h in enumerate(headers, start=1):
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if col_idx == 1:
                                cell.number_format = DATE_FORMAT
                            else:
                                cell.number_format = CURRENCY_FORMAT

                    # Ajuste automático de largura
                    for col_cells in ws.columns:
                        max_length = max(len(str(c.value)) for c in col_cells if c.value is not None)
                        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_length + 2

                    # Se excedeu parcelas e ainda há saldo devedor
                    if parcelas > 420 and saldo < 0:
                        st.error(
                            f"Financiamento de {cliente} não é possível! "
                            "A quantidade de parcelas excede 420 e o saldo devedor continua positivo."
                            f"Restariam {cliente} do saldo devedor."
                            "Simule novamente"
                            )
                    
                    # download
                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    st.download_button("Download Excel", data=buf,
                                    file_name=f"Financiamento {cliente}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            if __name__ == "__main__":
                main()
        else:
            st.error("Erro no login! Suas credenciais estão incorretas. " \
            "\nCaso não tenha uma credencial, entre em contato " \
            "\n com um gerente: (11)3047-2555")

def logout():
    if st.sidebar.button("Logout"):
        st.session_state.authenticated = False

# Check login status
if not st.session_state.authenticated:
    login()
else:
    st.sidebar.success(f"Logged in as {USERNAME}")
    logout()

    # 🔐 Protected content
    st.title("BR Financial Dashboard")
    st.write("Here is your secure content.")
