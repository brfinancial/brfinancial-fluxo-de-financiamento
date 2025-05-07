import streamlit as st
from io import BytesIO
import calendar
from datetime import datetime as dt, time
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Configurações fixas ---

TAXA_EMISSAO_CCB = 1500.0
TAXA_ALIENACAO_FIDUCIARIA = 2000.0
TAXA_REGISTRO_FIXA = 1500.0
TAXA_SEGURO_PRESTAMISTA_PCT = 0.083  # 8.3% pós-entrega
TAXA_INCC = 0.005  # 0.5% pré-entrega
TAXA_IPCA = 0.005  # 0.5% pós-entrega

HEADER_FILL = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
DATE_FORMAT = 'dd/mm/yyyy'
CURRENCY_FORMAT = '"R$" #.##0,00'
PERCENT_FORMAT = '0.00%'

# --- Auxiliares ---

def adjust_day(date, preferred_day):
    try:
        return date.replace(day=preferred_day)
    except ValueError:
        last = calendar.monthrange(date.year, date.month)[1]
        return date.replace(day=last)

def days_in_month(date):
    return calendar.monthrange(date.year, date.month)[1]

class PaymentTracker:
    def __init__(self, dia_pagamento, taxa_juros):
        self.last_date = None
        self.dia = dia_pagamento
        self.taxa = taxa_juros

    def calculate(self, current_date, saldo):
        dias_corridos = 0
        if self.last_date is not None:
            dias_corridos = (current_date - self.last_date).days
            taxa_efetiva = self.taxa * (dias_corridos / 30)
            juros = saldo * taxa_efetiva
        else:
            taxa_efetiva = 0.0
            juros = 0.0
        self.last_date = current_date
        return juros, dias_corridos, taxa_efetiva

# --- App Streamlit ---

def main():
    st.set_page_config(page_title="Gerador de Planilha de Financiamento", layout="centered")
    st.title("Bem-vindo ao gerador de financiamento da Br Financial!")

    # Entradas principais
    cliente = st.text_input("Qu9al o nome do cliente?")
    valor_imovel = st.number_input("Qual o valor total do imóvel (R$)", min_value=0.0, step=0.01, format="%.2f")
    dia_pagamento = st.number_input("Qual o dia preferencial de pagamento das parcelas mensais? (1-31)", min_value=1, max_value=31, step=1)
    taxa_pre = st.number_input("Taxa mensal de juros ANTES da entrega das chaves (%)", min_value=0.0, step=0.01) / 100
    taxa_pos = st.number_input("Taxa mensal de juros DEPOIS da entrega das chaves (%)", min_value=0.0, step=0.01) / 100

    # Data-base customizável
    data_base_date = st.date_input("Data-base (data de assinatura do contrato)", value=dt.now().date())
    base_date = dt.combine(data_base_date, time())

    # Taxas extras
    st.subheader("Taxas Extras")
    n_extras = st.number_input("Quantas taxas quer incluir nas parcelas? (Caso não tenha taxas extras, deixe em branco)", min_value=0, max_value=7, step=1)
    taxas_extras = []
    for i in range(int(n_extras)):
        pct = st.number_input(f"Taxa extra {i+1} (%)", min_value=0.0, step=0.01, key=f"pct_{i}") / 100
        periodo = st.selectbox(f"Período da taxa extra {i+1}", ["pré-entrega da chave", "pós-entrega da chave", "ambos"], key=f"periodo_{i}")
        taxas_extras.append({'pct': pct, 'periodo': periodo})

    # Datas e capacidades
    capacidade_pre = st.number_input("Qual a capacidade de pagamento do cliente nas parcelas mensais ANTES da entrega das chaves? (R$)", min_value=0.0, step=0.01)
    data_inicio_pre_date = st.date_input("Data início dos pagamentos mensais pré-entrega")
    data_entrega_date = st.date_input("Data de ENTREGA das chaves")
    data_inicio_pre = dt.combine(data_inicio_pre_date, time())
    data_entrega = dt.combine(data_entrega_date, time())
    fgts = st.number_input("Valor do FGTS para abatimento do saldo devedor (R$)", min_value=0.0, step=0.01)
    fin_banco = st.number_input("Valor financiado pelo banco (abatimento no saldo devedor) (R$)", min_value=0.0, step=0.01)
    capacidade_pos = st.number_input("Qual a capacidade de pagamento do cliente nas parcelas mensais DEPOIS da entrega das chaves? (R$)", min_value=0.0, step=0.01)

    # Pagamentos não recorrentes
    st.subheader("Pagamentos Não-Recorrentes")
    n_non_rec = st.number_input("Quantos pagamentos não recorrentes terão? (Caso não haja, deixe zerado)", min_value=0, step=1)
    non_rec = []
    for i in range(int(n_non_rec)):
        d_date = st.date_input(f"Data do pagamento {i+1}", key=f"nr_d_{i}")
        d = dt.combine(d_date, time())
        v = st.number_input(f"Valor pagamento {i+1} (R$)", min_value=0.0, step=0.01, key=f"nr_v_{i}")
        desc = st.text_input(f"Descrição do pagamento {i+1}", key=f"nr_desc_{i}")
        assoc = st.checkbox(f"Atribuir a parcela normal do mês? {i+1}", key=f"nr_assoc_{i}")
        if assoc:
            d = adjust_day(d, dia_pagamento)
        non_rec.append({'data': d, 'tipo': desc, 'valor': v, 'assoc': assoc})

    # Séries semestrais e anuais
    st.subheader("Pagamentos Semestrais Recorrentes")
    n_semi = st.number_input("Quantos pagamentos recorrentes semestrais terão? (Caso não haja, deixe zerado)", min_value=0, step=1)
    semi_series = []
    for i in range(int(n_semi)):
        d0_date = st.date_input(f"Data das parcelas semestrais {i+1}", key=f"s_d0_{i}")
        d0 = dt.combine(d0_date, time())
        v = st.number_input(f"Valor da parcela semestral {i+1} (R$)", min_value=0.0, step=0.01, key=f"s_v_{i}")
        assoc = st.checkbox(f"Atribuir a parcela normal do mês? {i+1}", key=f"s_assoc_{i}")
        semi_series.append({'d0': d0, 'v': v, 'assoc': assoc, 'tipo': 'Pagamento Semestral'})

    st.subheader("Pagamentos Anuais Recorrentes")
    n_ann = st.number_input("Quantos pagamentos recorrentes anuais terão? (Caso não haja, deixe zerado)", min_value=0, step=1)
    annual_series = []
    for i in range(int(n_ann)):
        d0_date = st.date_input(f"Data das parcelas anuais {i+1}", key=f"a_d0_{i}")
        d0 = dt.combine(d0_date, time())
        v = st.number_input(f"Valor da parcela anual {i+1} (R$)", min_value=0.0, step=0.01, key=f"a_v_{i}")
        assoc = st.checkbox(f"Atribuir a parcela normal do mês? {i+1}", key=f"a_assoc_{i}")
        annual_series.append({'d0': d0, 'v': v, 'assoc': assoc, 'tipo': 'Pagamento Anual'})

    # Geração da planilha
    if st.button("Gerar Planilha"):
        # --- Agrega séries em non_rec ---
        for series in semi_series:
            for n in range(100):
                d = series['d0'] + relativedelta(months=6 * n)
                assoc = series['assoc']
                if assoc:
                    d = adjust_day(d, dia_pagamento)
                non_rec.append({'data': d, 'tipo': series['tipo'], 'valor': series['v'], 'assoc': assoc})
        for series in annual_series:
            for n in range(100):
                d = series['d0'] + relativedelta(years=n)
                assoc = series['assoc']
                if assoc:
                    d = adjust_day(d, dia_pagamento)
                non_rec.append({'data': d, 'tipo': series['tipo'], 'valor': series['v'], 'assoc': assoc})

        # --- Separa pré e pós entre non_rec ---
        pre_nr = sorted([e for e in non_rec if e['data'] < data_entrega], key=lambda x: x['data'])
        post_nr = sorted([e for e in non_rec if e['data'] >= data_entrega], key=lambda x: x['data'])

        eventos = []
        saldo = valor_imovel

        # Data base (assinatura do contrato)
        eventos.append({
            'data': base_date,
            'parcela': '',
            'tipo': 'Data-Base (assinatura do contrato)',
            'valor': 0.0,
            'juros': 0.0,
            'dias_corridos': 0,
            'taxa_efetiva': 0.0,
            'incc': 0.0,
            'ipca': 0.0,
            'taxas_extra': [0.0] * len(taxas_extras),
            'Total de mudança (R$)': 0.0,
            'saldo': saldo
        })

        tracker_pre = PaymentTracker(dia_pagamento, taxa_pre)
        tracker_pre.last_date = base_date

        # 1) PRÉ-ENTREGA ------------------------------------------------
        prev_date = data_inicio_pre
        cursor = data_inicio_pre
        last_pre_date = None
        while True:
            d_evt = adjust_day(cursor, dia_pagamento)
            if d_evt >= data_entrega:
                last_pre_date = prev_date
                break
            # não-recorrentes pré não associados entre prev_date e d_evt
            for ev_nr in [e for e in pre_nr if not e['assoc'] and prev_date < e['data'] < d_evt]:
                juros, dias_corr, taxa_eff = tracker_pre.calculate(ev_nr['data'], saldo)
                incc_nr = saldo * TAXA_INCC
                extras_nr = [saldo * t['pct'] if t['periodo'] in ['pré-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                total_taxas_nr = sum(extras_nr) + incc_nr
                abat_nr = ev_nr['valor'] - juros - total_taxas_nr
                saldo -= abat_nr
                eventos.append({**ev_nr, 'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                                'incc': incc_nr, 'ipca': 0.0, 'taxas_extra': extras_nr,
                                'Total de mudança (R$)': abat_nr, 'saldo': saldo})
            # Parcela mensal pré-entrega com associados
            assoc_events = [e for e in pre_nr if e['assoc'] and e['data'] == d_evt]
            extra_val = sum(e['valor'] for e in assoc_events)
            types_concat = ' + '.join(['Pré-Entrega'] + [e['tipo'] for e in assoc_events]) if assoc_events else 'Pré-Entrega'
            juros, dias_corr, taxa_eff = tracker_pre.calculate(d_evt, saldo)
            incc = saldo * TAXA_INCC
            extras = [saldo * t['pct'] if t['periodo'] in ['pré-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
            total_taxas = sum(extras) + incc
            total_payment = capacidade_pre + extra_val
            abat = total_payment - juros - total_taxas
            if abat <= 0:
                st.error(f"O abatimento na parcela {parcelas} é zero ou negativo. Financiamento não viável.")
                break
            saldo -= abat
            eventos.append({'data': d_evt, 'parcela': '', 'tipo': types_concat, 'valor': total_payment,
                            'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                            'incc': incc, 'ipca': 0.0, 'taxas_extra': extras,
                            'Total de mudança (R$)': abat, 'saldo': saldo})
            prev_date = d_evt
            cursor += relativedelta(months=1)

        # 2) ENTREGA ------------------------------------------------------
        ent = adjust_day(data_entrega, dia_pagamento)
        zero_extras = [0.0] * len(taxas_extras)
        for desc, v in [('Abatimento FGTS', fgts), ('Abatimento Fin. Banco', fin_banco)]:
            saldo -= v
            eventos.append({'data': ent, 'parcela': '', 'tipo': desc, 'valor': 0.0,
                            'juros': 0.0, 'dias_corridos': '', 'taxa_efetiva': '',
                            'incc': 0.0, 'ipca': 0.0, 'taxas_extra': zero_extras,
                            'Total de mudança (R$)': v, 'saldo': saldo})
        for nome, val in [('Emissão CCB', TAXA_EMISSAO_CCB), ('Alienação Fiduciária', TAXA_ALIENACAO_FIDUCIARIA),
                          ('Registro', TAXA_REGISTRO_FIXA)]:
            saldo += val
            eventos.append({'data': ent, 'parcela': '', 'tipo': 'Taxa ' + nome, 'valor': 0.0,
                            'juros': 0.0, 'dias_corridos': '', 'taxa_efetiva': '',
                            'incc': 0.0, 'ipca': 0.0, 'taxas_extra': zero_extras,
                            'Total de mudança (R$)': val, 'saldo': saldo})
        fee = saldo * TAXA_SEGURO_PRESTAMISTA_PCT
        saldo += fee
        eventos.append({'data': ent, 'parcela': '', 'tipo': 'Taxa Seguro Prestamista', 'valor': 0.0,
                        'juros': 0.0, 'dias_corridos': '', 'taxa_efetiva': '',
                        'incc': 0.0, 'ipca': 0.0, 'taxas_extra': zero_extras,
                        'Total de mudança (R$)': fee, 'saldo': saldo})

        # ao término da pré-entrega, capture exatamente a data do último evento:
        last_pre_event = sorted([e['data'] for e in eventos if e['data'] < data_entrega])[-1]
        
        # na inicialização do tracker pós-entrega, use last_pre_event:
        tracker_pos = PaymentTracker(dia_pagamento, taxa_pos)
        tracker_pos.last_date = last_pre_event

        # fluxo pós-entrega (permanece igual, mas com dias_corridos corretos)
        prev_date = data_entrega
        cursor = data_entrega
        parcelas = 1
        while saldo > 0 and parcelas <= 420:
            d_evt = adjust_day(cursor + relativedelta(months=1), dia_pagamento)

        # 3) PÓS-ENTREGA --------------------------------------------------
        tracker_pos = PaymentTracker(dia_pagamento, taxa_pos)
        tracker_pos.last_date = last_pre_date
        prev_date = data_entrega
        cursor = data_entrega
        parcelas = 1
        while saldo > 0 and parcelas <= 420:
            d_evt = adjust_day(cursor + relativedelta(months=1), dia_pagamento)
            # não-recorrentes pós não associados entre prev_date e d_evt
            for ev_nr in [e for e in post_nr if not e['assoc'] and prev_date < e['data'] < d_evt]:
                juros, dias_corr, taxa_eff = tracker_pos.calculate(ev_nr['data'], saldo)
                ipca_nr = saldo * TAXA_IPCA
                extras_nr = [saldo * t['pct'] if t['periodo'] in ['pós-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                total_taxas_nr = sum(extras_nr) + ipca_nr
                abat_nr = ev_nr['valor'] - juros - total_taxas_nr
                saldo -= abat_nr
                eventos.append({**ev_nr,'parcela': parcelas, 'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                                'incc': 0.0, 'ipca': ipca_nr, 'taxas_extra': extras_nr,
                                'Total de mudança (R$)': abat_nr, 'saldo': saldo})
            # parcela mensal pós-entrega com associados
            assoc_events = [e for e in post_nr if e['assoc'] and e['data'] == d_evt]
            extra_val = sum(e['valor'] for e in assoc_events)
            types_concat = ' + '.join(['Pós-Entrega'] + [e['tipo'] for e in assoc_events]) if assoc_events else 'Pós-Entrega'
            juros, dias_corr, taxa_eff = tracker_pos.calculate(d_evt, saldo)
            ipca = saldo * TAXA_IPCA
            extras = [saldo * t['pct'] if t['periodo'] in ['pós-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
            total_taxas = sum(extras) + ipca
            total_payment = capacidade_pos + extra_val
            abat = total_payment - juros - total_taxas
            saldo -= abat
            eventos.append({'data': d_evt, 'parcela': parcelas, 'tipo': types_concat, 'valor': total_payment,
                            'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                            'incc': 0.0, 'ipca': ipca, 'taxas_extra': extras,
                            'Total de mudança (R$)': abat, 'saldo': saldo})
            parcelas += 1
            prev_date = d_evt
            cursor = d_evt

        if parcelas >= 420 and saldo > 0:
            st.error(f"Financiamento de {cliente} não é possível! A quantidade de parcelas excede 420 e o saldo devedor continua positivo.")

        # --- Montar planilha ---
        wb = Workbook()
        ws = wb.active
        ws.title = f"Financ-{cliente}"[:31]
        headers = ["Data","Parcela","Tipo","Dias no Mês","Dias Corridos","Taxa Efetiva","Valor Pago (R$)",
                   "Juros (R$)","INCC (R$)","IPCA (R$)"]
        headers += [f"Taxa {i+1} (R$)" for i in range(len(taxas_extras))]
        headers += ["Total de adições e subtrações (R$)","Saldo Devedor (R$)"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
        # linha inicial
        ws.append(["-"]*(len(headers)-1) + [valor_imovel])
        # eventos
        for ev in sorted(eventos, key=lambda x: x['data']):
            row = [ev['data'], ev.get('parcela', ''), ev['tipo'], days_in_month(ev['data']),
                   ev.get('dias_corridos', ''), ev.get('taxa_efetiva', ''), ev.get('valor', 0),
                   ev.get('juros', 0), ev.get('incc', 0), ev.get('ipca', 0)]
            row += ev.get('taxas_extra', []) + [ev.get('Total de mudança (R$)', 0), ev.get('saldo', 0)]
            ws.append(row)

        # 7) Ajuste automático de largura das colunas
        for col_cells in ws.columns:
            max_length = 0
            column = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            ws.column_dimensions[column].width = max_length + 2  # **Reinserido: ajusta largura automática**

        # 8) Formatação: Data, inteiro, porcentagem e moeda (com R$)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if header == "Data":
                    cell.number_format = DATE_FORMAT
                elif header in ["Parcela", "Dias no Mês", "Dias Corridos"]:
                    cell.number_format = '0'
                elif header == "Taxa Efetiva":
                    cell.number_format = PERCENT_FORMAT  # **Reinserido: '%'**
                else:
                    cell.number_format = CURRENCY_FORMAT  # **Reinserido: 'R$'**

        # Totais e formatação (como antes) ...
        # [mantido sem alteração]        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button("Download Excel", data=buf,
                           file_name=f"financiamento_{cliente}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    main()
